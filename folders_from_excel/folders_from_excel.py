from openpyxl import Workbook as wb
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
import os

from openpyxl.reader import excel


def get_file_path():
    """
    This function obtains Excel file path using tkinter module.
    """
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Excel file", "*.xlsx")])
    return file_path


def read_excel_column(file_path):
    """
    This function uses openpyxl module to read all rows of first column of chosen Excel file.
    It then writes rows' values to a list.
    """
    excel = load_workbook(file_path)
    ws = excel.active

    folder_names = [ws[f"A{row}"].value for row in range(1, ws.max_row + 1)]

    return folder_names


def get_directory():
    """
    This function uses tkinter to get output directory path from user.
    """
    root = tk.Tk()
    root.withdraw()

    directory = filedialog.askdirectory()
    return directory


def make_folders(directory, folder_names):
    """
    This function creates directories using Python os module.
    Path of a directory is created using chosen directory name and values of excel sheet cells.
    """

    for folder in folder_names:
        if directory == "":
            path = folder
        else:
            path = directory + "/" + folder
        if not os.path.exists(path):
            os.mkdir(path)
        else:
            print("Directory", path, "already exists")


if get_file_path() == "":
    file_path = get_file_path()
    folder_names = read_excel_column(file_path)
    directory = get_directory()
    make_folders(directory, folder_names)
